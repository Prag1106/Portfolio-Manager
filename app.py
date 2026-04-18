from flask import Flask, render_template, request, jsonify, send_file
import sqlite3, json, io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference

app = Flask(__name__)
DB = 'portfolio.db'

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as db:
        db.execute('''CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        db.execute('''CREATE TABLE IF NOT EXISTS portfolios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            type TEXT NOT NULL,
            data TEXT NOT NULL,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE
        )''')
        db.commit()

init_db()

@app.route('/')
def index():
    return render_template('index.html')

# ── Clients ───────────────────────────────────────────────────────────────────
@app.route('/api/clients', methods=['GET'])
def get_clients():
    db = get_db()
    clients = db.execute('SELECT id,name,created_at,updated_at FROM clients ORDER BY name').fetchall()
    result = []
    for c in clients:
        portfolios = db.execute('SELECT type FROM portfolios WHERE client_id=?',(c['id'],)).fetchall()
        result.append({
            'id':c['id'],'name':c['name'],
            'created_at':c['created_at'],'updated_at':c['updated_at'],
            'portfolios':[p['type'] for p in portfolios]
        })
    return jsonify(result)

@app.route('/api/clients', methods=['POST'])
def create_client():
    data = request.json
    name = data.get('name','').strip()
    if not name: return jsonify({'error':'Name required'}),400
    db = get_db()
    cur = db.execute('INSERT INTO clients (name) VALUES (?)',(name,))
    db.commit()
    return jsonify({'id':cur.lastrowid,'name':name})

@app.route('/api/clients/<int:cid>', methods=['PUT'])
def update_client(cid):
    data = request.json
    name = data.get('name','').strip()
    if not name: return jsonify({'error':'Name required'}),400
    db = get_db()
    db.execute('UPDATE clients SET name=?,updated_at=? WHERE id=?',(name,datetime.now().isoformat(),cid))
    db.commit()
    return jsonify({'ok':True})

@app.route('/api/clients/<int:cid>', methods=['DELETE'])
def delete_client(cid):
    db = get_db()
    db.execute('DELETE FROM clients WHERE id=?',(cid,))
    db.commit()
    return jsonify({'ok':True})

@app.route('/api/clients/<int:cid>/duplicate', methods=['POST'])
def duplicate_client(cid):
    db = get_db()
    orig = db.execute('SELECT name FROM clients WHERE id=?',(cid,)).fetchone()
    if not orig: return jsonify({'error':'Not found'}),404
    new_name = request.json.get('name', orig['name']+' (copia)').strip()
    cur = db.execute('INSERT INTO clients (name) VALUES (?)',(new_name,))
    new_id = cur.lastrowid
    portfolios = db.execute('SELECT type,data FROM portfolios WHERE client_id=?',(cid,)).fetchall()
    now = datetime.now().isoformat()
    for p in portfolios:
        db.execute('INSERT INTO portfolios (client_id,type,data,updated_at) VALUES (?,?,?,?)',
                   (new_id,p['type'],p['data'],now))
    db.commit()
    return jsonify({'id':new_id,'name':new_name})

# ── Portfolios ────────────────────────────────────────────────────────────────
@app.route('/api/clients/<int:cid>/portfolios', methods=['GET'])
def get_portfolios(cid):
    db = get_db()
    rows = db.execute('SELECT type,data,updated_at FROM portfolios WHERE client_id=?',(cid,)).fetchall()
    return jsonify({r['type']:{'data':json.loads(r['data']),'updated_at':r['updated_at']} for r in rows})

@app.route('/api/clients/<int:cid>/portfolios', methods=['POST'])
def save_portfolios(cid):
    data = request.json
    db = get_db()
    now = datetime.now().isoformat()
    existing = {r['type'] for r in db.execute('SELECT type FROM portfolios WHERE client_id=?',(cid,)).fetchall()}
    for ptype,pdata in data.items():
        if ptype in existing:
            db.execute('UPDATE portfolios SET data=?,updated_at=? WHERE client_id=? AND type=?',
                       (json.dumps(pdata),now,cid,ptype))
        else:
            db.execute('INSERT INTO portfolios (client_id,type,data,updated_at) VALUES (?,?,?,?)',
                       (cid,ptype,json.dumps(pdata),now))
    for t in existing:
        if t not in data:
            db.execute('DELETE FROM portfolios WHERE client_id=? AND type=?',(cid,t))
    db.execute('UPDATE clients SET updated_at=? WHERE id=?',(now,cid))
    db.commit()
    return jsonify({'ok':True})

# ── ISIN lookup proxy ─────────────────────────────────────────────────────────
@app.route('/api/isin/<isin>', methods=['GET'])
def lookup_isin(isin):
    import urllib.request, urllib.parse
    isin = isin.strip().upper()
    euro = ['MIL','GER','FRA','AMS','LSE','EBS','VIE','BRU','CPH','HEL','OSL','STK','MCE','SWX']
    hdrs = {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120','Accept':'application/json'}
    quotes = []
    for host in ['query1','query2']:
        try:
            url = f'https://{host}.finance.yahoo.com/v1/finance/search?q={urllib.parse.quote(isin)}&quotesCount=10&newsCount=0&listsCount=0&enableFuzzyQuery=false'
            req = urllib.request.Request(url, headers=hdrs)
            with urllib.request.urlopen(req, timeout=10) as r:
                quotes = json.loads(r.read()).get('quotes',[])
            if quotes: break
        except: continue
    if not quotes: return jsonify({'error':'not found'}),404
    best = (next((q for q in quotes if q.get('exchange') in euro and q.get('quoteType') in ('ETF','MUTUALFUND')),None)
         or next((q for q in quotes if q.get('quoteType') in ('ETF','MUTUALFUND')),None)
         or next((q for q in quotes if q.get('exchange') in euro),None)
         or quotes[0])
    name = best.get('longname') or best.get('shortname') or best.get('symbol',isin)
    qt = best.get('quoteType','')
    ftype = 'ETF' if 'ETF' in qt else 'Azione' if 'EQUITY' in qt else 'Obbligazione' if any(x in qt for x in ['BOND','INCOME']) else 'Fondo'
    symbol = best.get('symbol','')
    price = None
    if symbol:
        for host in ['query1','query2']:
            try:
                url = f'https://{host}.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}?interval=1d&range=1d'
                req = urllib.request.Request(url, headers=hdrs)
                with urllib.request.urlopen(req, timeout=10) as r:
                    pd = json.loads(r.read())
                meta = pd.get('chart',{}).get('result',[{}])[0].get('meta',{})
                price = meta.get('regularMarketPrice') or meta.get('previousClose')
                if price: break
            except: continue
    return jsonify({'name':name,'price':price,'type':ftype,'ticker':symbol})

# ── Excel export ──────────────────────────────────────────────────────────────
@app.route('/api/clients/<int:cid>/export', methods=['GET'])
def export_excel(cid):
    db = get_db()
    client = db.execute('SELECT name FROM clients WHERE id=?',(cid,)).fetchone()
    if not client: return jsonify({'error':'Not found'}),404
    rows = db.execute('SELECT type,data FROM portfolios WHERE client_id=?',(cid,)).fetchall()

    wb = Workbook()
    wb.remove(wb.active)

    NAVY='0F1E35'; GOLD='C9A84C'; GOLD_L='E8C97A'; WHITE='FFFFFF'
    EQ_P='EAF7F0'; EQ_D='1A6E3C'; EQ_M='2EAA5F'
    BD_P='EAF1FB'; BD_D='1A4A8A'; BD_M='2E72D2'
    CM_P='FEF6E8'; CM_D='8A4A00'; CM_M='D97F00'
    SG_P='FDEEF4'; SG_D='7A1A3C'; SG_M='C03060'
    BGREY='F1F3F8'; LGREY='F8F9FB'; BORDER='E0E4EC'
    AC_COLORS = {
        'EQUITY':(EQ_P,EQ_D,EQ_M),'BOND':(BD_P,BD_D,BD_M),
        'COMMODITY':(CM_P,CM_D,CM_M),'START & GO':(SG_P,SG_D,SG_M)
    }
    DEF_COLOR = ('F0F2F5','2C3E50','7F8C8D')

    def fl(c): return PatternFill('solid',fgColor=c)
    def fn(bold=False,color='1A1F2E',size=10,italic=False,mono=False):
        return Font(name='Courier New' if mono else 'Calibri',bold=bold,color=color,size=size,italic=italic)
    def al(h='left',v='center'): return Alignment(horizontal=h,vertical=v,wrap_text=False)
    thin = Side(style='thin',color=BORDER)
    def bdr(): return Border(left=thin,right=thin,top=thin,bottom=thin)

    TL = {'amm':'Amministrato','darta':'DARTA','pac':'PAC','sicav':'SICAV & Fondi'}

    for row in rows:
        pt = row['type']
        pd_data = json.loads(row['data'])
        label = TL.get(pt,pt.upper())
        ws = wb.create_sheet(title=label)
        ws.sheet_view.showGridLines = False

        for c,w in {1:3,2:16,3:34,4:13,5:12,6:14,7:12,8:12,9:12}.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        r = 1
        # Title — write to each cell individually, no merge
        ws.row_dimensions[r].height = 36
        title_text = f'PORTFOLIO REPORT — {client["name"].upper()} — {label.upper()}'
        ws.cell(r,2,title_text).font = fn(True,WHITE,13)
        ws.cell(r,2).fill = fl(NAVY); ws.cell(r,2).alignment = al('center')
        for col in range(2,10): ws.cell(r,col).fill = fl(NAVY)
        r += 1

        ws.row_dimensions[r].height = 16
        sub_text = f'Data: {datetime.now().strftime("%d/%m/%Y")}   |   Totale investito: €{pd_data.get("total",0):,.0f}'
        ws.cell(r,2,sub_text).font = fn(False,GOLD_L,9,True)
        ws.cell(r,2).fill = fl(NAVY); ws.cell(r,2).alignment = al('center')
        for col in range(2,10): ws.cell(r,col).fill = fl(NAVY)
        r += 2

        weights = pd_data.get('weights',{})
        fund_rows = pd_data.get('rows',{})
        total = pd_data.get('total',0)
        chart_labels = []; chart_vals = []

        # Preserve order using list
        weight_items = list(weights.items())

        for ac,acd in weight_items:
            if ac == 'START & GO': continue
            ac_pct = acd.get('pct',0)/100
            ac_val = total*ac_pct
            pale,dark,mid = AC_COLORS.get(ac,DEF_COLOR)
            chart_labels.append(ac); chart_vals.append(round(ac_pct*100,1))

            # AC header — single cell only, no merge
            ws.row_dimensions[r].height = 22
            ws.cell(r,2,f'  {ac}').font = fn(True,dark,12)
            ws.cell(r,2).fill = fl(pale)
            ws.cell(r,2).alignment = al()
            ws.cell(r,2).border = Border(top=Side(style='medium',color=mid))
            ws.cell(r,7,f'€{ac_val:,.0f}').font = fn(True,dark,11)
            ws.cell(r,7).fill = fl(pale); ws.cell(r,7).alignment = al('right')
            ws.cell(r,9,f'{ac_pct*100:.1f}%').font = fn(True,dark,11)
            ws.cell(r,9).fill = fl(pale); ws.cell(r,9).alignment = al('right')
            for col in range(2,10):
                ws.cell(r,col).fill = fl(pale)
                ws.cell(r,col).border = Border(top=Side(style='medium',color=mid))
            r += 1

            # Column headers
            ws.row_dimensions[r].height = 15
            hdrs = ['','ISIN','Nome Fondo','Prezzo','% Settore','CTRV','Quantità','% Totale','% Effettivo']
            for ci,h in enumerate(hdrs,1):
                c = ws.cell(r,ci,h)
                c.font = fn(True,WHITE,9); c.fill = fl(NAVY)
                c.alignment = al('center'); c.border = bdr()
            r += 1

            subs = list(acd.get('subs',{}).items())
            for sub,sp100 in subs:
                sub_pct = sp100/100; obj_pct = ac_pct*sub_pct; sv = total*obj_pct

                # Sub header — single cell
                ws.row_dimensions[r].height = 18
                ws.cell(r,2,f'  ▸  {sub.upper()}').font = fn(True,'1A1F2E',11)
                ws.cell(r,2).fill = fl(BGREY); ws.cell(r,2).alignment = al()
                ws.cell(r,4,f'€{sv:,.0f}').font = fn(True,'4A5168',10)
                ws.cell(r,4).fill = fl(BGREY); ws.cell(r,4).alignment = al('right')
                ws.cell(r,5,f'{obj_pct*100:.1f}%').font = fn(False,'8892A8',9,True)
                ws.cell(r,5).fill = fl(BGREY)
                for col in range(2,10):
                    ws.cell(r,col).fill = fl(BGREY); ws.cell(r,col).border = bdr()
                r += 1

                types = ['Obbligazione','ETF'] if ac=='BOND' else ['Azione','ETF']
                for ftype in types:
                    ws.row_dimensions[r].height = 13
                    ws.cell(r,2,f'    {ftype}').font = fn(False,'8892A8',9,True)
                    ws.cell(r,2).fill = fl('FAFBFD')
                    for col in range(2,10): ws.cell(r,col).fill = fl('FAFBFD')
                    r += 1

                    # Get all rows for this type
                    type_keys = sorted(
                        [k for k in fund_rows if k.startswith(f'{pt}|{ac}|{sub}|{ftype}|')],
                        key=lambda k: int(k.split('|')[4])
                    )
                    if not type_keys:
                        type_keys = [f'{pt}|{ac}|{sub}|{ftype}|0']

                    for k in type_keys:
                        rd = fund_rows.get(k,{})
                        isin=rd.get('isin',''); nm=rd.get('name','')
                        price=rd.get('price'); ps=rd.get('pctSettore',0)
                        ctrv=sv*ps if ps else 0
                        qty=ctrv/price if (price and price>0 and ctrv>0) else None

                        ws.row_dimensions[r].height = 15
                        bg = WHITE if (r%2==0) else 'FAFBFC'
                        vals = ['',isin,nm,
                                f'€{price:.2f}' if price else '—',
                                f'{ps*100:.1f}%' if ps else '—',
                                f'€{ctrv:,.0f}' if ctrv else '—',
                                f'{qty:,.3f}' if qty else '—',
                                f'{(ctrv/total)*100:.2f}%' if (ctrv and total) else '—',
                                '—']
                        for ci,v in enumerate(vals,1):
                            c = ws.cell(r,ci,v)
                            c.fill = fl(bg); c.border = bdr()
                            c.alignment = al('right' if ci>3 else 'left')
                            if ci==2: c.font = fn(False,BD_D,9,mono=True)
                            elif ci==3: c.font = fn(bool(nm),'1A1F2E',9)
                            else: c.font = fn(False,'1A1F2E' if v!='—' else 'C8CDD8',9)
                        r += 1

                    # Sub total
                    sub_ctrv = sum(
                        sv * fund_rows.get(k,{}).get('pctSettore',0)
                        for k in fund_rows if k.startswith(f'{pt}|{ac}|{sub}|')
                    )
                    ws.row_dimensions[r].height = 14
                    ws.cell(r,2,f'Totale {sub}').font = fn(True,'4A5168',9)
                    ws.cell(r,2).fill = fl('EEF0F4')
                    ws.cell(r,6,f'€{sub_ctrv:,.0f}' if sub_ctrv else '—').font = fn(True,'1A1F2E',9)
                    ws.cell(r,6).fill = fl('EEF0F4'); ws.cell(r,6).alignment = al('right')
                    ws.cell(r,8,f'{(sub_ctrv/total)*100:.2f}%' if (sub_ctrv and total) else '—').font = fn(True,'1A1F2E',9)
                    ws.cell(r,8).fill = fl('EEF0F4'); ws.cell(r,8).alignment = al('right')
                    for col in range(2,10):
                        ws.cell(r,col).fill = fl('EEF0F4'); ws.cell(r,col).border = bdr()
                    r += 1
            r += 1

        # Grand total — single cell
        grand = sum(
            total*(acd.get('pct',0)/100)*(sp100/100)*
            fund_rows.get(k,{}).get('pctSettore',0)
            for ac,acd in weight_items if ac!='START & GO'
            for sub,sp100 in acd.get('subs',{}).items()
            for k in fund_rows if k.startswith(f'{pt}|{ac}|{sub}|')
        )
        ws.row_dimensions[r].height = 24
        ws.cell(r,2,'TOTALE PORTAFOGLIO').font = fn(True,WHITE,11)
        ws.cell(r,2).fill = fl(NAVY); ws.cell(r,2).alignment = al()
        ws.cell(r,6,f'€{grand:,.0f}').font = fn(True,GOLD_L,12)
        ws.cell(r,6).fill = fl(NAVY); ws.cell(r,6).alignment = al('right')
        for col in range(2,10): ws.cell(r,col).fill = fl(NAVY)
        r += 3

        # Pie chart
        if chart_labels:
            cr = r
            ws.cell(r,2,'Asset Class').font = fn(True,'8892A8',9)
            ws.cell(r,3,'Peso %').font = fn(True,'8892A8',9)
            r += 1
            for lbl,val in zip(chart_labels,chart_vals):
                ws.cell(r,2,lbl).font = fn(False,'1A1F2E',9)
                c3 = ws.cell(r,3,val)
                c3.number_format = '0.0"%"'
                r += 1
            pie = PieChart(); pie.title='Asset Allocation'; pie.style=10
            pie.width=14; pie.height=10
            pie.add_data(Reference(ws,min_col=3,min_row=cr,max_row=cr+len(chart_labels)),titles_from_data=True)
            pie.set_categories(Reference(ws,min_col=2,min_row=cr+1,max_row=cr+len(chart_labels)))
            ws.add_chart(pie,f'E{cr}')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"{client['name'].replace(' ','_')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

import os
if __name__=='__main__':
    port=int(os.environ.get('PORT',5001))
    debug=os.environ.get('RAILWAY_ENVIRONMENT') is None
    app.run(debug=debug,host='0.0.0.0',port=port)
